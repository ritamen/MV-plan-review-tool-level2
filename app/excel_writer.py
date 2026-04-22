"""
excel_writer.py
---------------
All openpyxl logic for writing AI review results into the M&V Plan Review Sheet.
The AI never sees or touches this file — Python handles everything.

Sheet  : "1. M&V plan_V2.0"
Columns written (1-based):
  H  col 8  = Included
  I  col 9  = Active Status
  J  col 10 = Consultant's Comments Round 1

Colour coding:
  Yes / APP -> bg #00B050  font #FFFFFF  bold  centered
  Yes / AAN -> bg #00B0F0  font #FFFFFF  bold  centered
  No  / NA  -> bg #FF0000  font #FFFFFF  bold  centered
  Partial / IR -> bg #FFFF00  font #000000  bold  centered

  Comment non-empty   -> bg #FFFF99  font #000000  wrap  top-aligned
  Comment empty       -> bg #FFFFFF
  Row height          -> max(40, (len//80 + newlines + 1) * 15)

  All written cells: Trebuchet MS 10, thin border #BBBBBB.
"""

import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ── Column indices (1-based) ─────────────────────────────────────────────────
COL_SN      = 2   # B
COL_INCL    = 8   # H  Included
COL_STATUS  = 9   # I  Active Status
COL_COMMENT = 10  # J  Consultant's Comments Round 1
DATA_START  = 22  # First data row

FONT_NAME = "Trebuchet MS"
FONT_SIZE = 10
BORDER_COLOR = "BBBBBB"

# ── Style maps ───────────────────────────────────────────────────────────────
INCLUDED_STYLES = {
    "Yes":     {"bg": "E2EFDA", "fc": "00B050"},
    "No":      {"bg": "FFD3D3", "fc": "C00000"},
    "Partial": {"bg": "FFF2CC", "fc": "C65911"},
}

STATUS_STYLES = {
    "APP": {"bg": "00B050", "fc": "FFFFFF"},
    "AAN": {"bg": "00B0F0", "fc": "FFFFFF"},
    "NA":  {"bg": "FF0000", "fc": "FFFFFF"},
    "IR":  {"bg": "FFFF00", "fc": "000000"},
}

COMMENT_BG = "FFFF99"
COMMENT_FC = "000000"

# ── Calc/Report sheet style maps — same colours and font as Sheet 1 ──────────
CALC_FONT_SIZE = FONT_SIZE   # Trebuchet MS 10, consistent with Sheet 1
CALC_BORDER_COLOR = "BFBFBF"

CALC_INCLUDED_STYLES = INCLUDED_STYLES   # identical to Sheet 1

CALC_STATUS_STYLES = STATUS_STYLES       # identical to Sheet 1 (APP/AAN/NA/IR)


def _normalize_table_borders(ws, start_row: int = DATA_START) -> None:
    """
    Apply uniform thin borders to every non-slave cell in the table area
    (col B to max_column, start_row to max_row). Leaves cell values and
    fill colors untouched — only the border style is overwritten.
    """
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Collect slave cells of merged ranges so we can skip them
    slave_cells: set[tuple[int, int]] = set()
    for merge in ws.merged_cells.ranges:
        for r in range(merge.min_row, merge.max_row + 1):
            for c in range(merge.min_col, merge.max_col + 1):
                if r != merge.min_row or c != merge.min_col:
                    slave_cells.add((r, c))

    max_col = ws.max_column or 17
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in range(2, max_col + 1):
            if (row_idx, col_idx) in slave_cells:
                continue
            ws.cell(row=row_idx, column=col_idx).border = border


def _sheet_assessment(statuses: list) -> str:
    """
    Compute overall sheet assessment from a list of question statuses.
      APP — every question is APP
      IR  — majority (> 50 %) are IR
      NA  — majority (> 50 %) are NA, or tied IR/NA (NA is more severe)
    AAN is not used in Round 1.
    """
    if not statuses:
        return "NA"
    total = len(statuses)
    n_app = statuses.count("APP")
    n_ir  = statuses.count("IR")
    n_na  = statuses.count("NA")
    if n_app == total:
        return "APP"
    if n_ir > total / 2:
        return "IR"
    if n_na > total / 2:
        return "NA"
    # tied or mixed — pick the more severe
    return "NA" if n_na >= n_ir else "IR"


def _make_border() -> Border:
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _style_cell(cell, value: str, bg: str, fc: str,
                bold: bool = False, wrap: bool = False,
                align: str = "left") -> None:
    cell.value = value
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, color=fc, bold=bold)
    cell.fill = _make_fill(bg)
    vertical = "center" if align == "center" else "top"
    cell.alignment = Alignment(horizontal=align, vertical=vertical, wrap_text=wrap)
    cell.border = _make_border()


def _is_section_header(value) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    if not s:
        return True
    try:
        f = float(s)
        return f == int(f) and "." not in s
    except (ValueError, TypeError):
        return False  # unparseable as float → multi-part SN like "6.3.1", treat as question row


def _is_integer_section(value) -> bool:
    """True only for whole-integer section numbers like 0, 1, 2 … 17 (not empty/None)."""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    try:
        f = float(s)
        return f == int(f) and "." not in s
    except (ValueError, TypeError):
        return False


def _write_section_rollups(ws, review_by_sn: dict) -> None:
    """
    Write rollup Included / Active Status to every section / sub-section header
    row (integer headers like 1, 2 … AND sub-section headers like 6.3, 6.4)
    based solely on the statuses of the question rows beneath them.
    No AI involvement — purely computed from sub-question results.

    A row is treated as a rollup header if:
      • Its SN is a whole integer (0, 1, 2 … 17), OR
      • Any other SN in the sheet starts with that SN + "." (e.g. 6.3 because
        6.3.1 exists).

    Rollup rules (same for all header types):
      APP     — every sub-question is APP
      NA      — majority (> 50 %) of sub-questions are NA
      IR      — majority (> 50 %) of sub-questions are IR
      IR      — mixed / tied (no clear majority)

    Included mirrors status: Yes → APP, No → NA, Partial → IR.
    No comment is written to section header rows.
    """
    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # Pre-scan: collect every non-blank SN string from column B
    sn_positions: list[tuple[int, str]] = []
    for row_idx in range(DATA_START, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=COL_SN).value
        if raw is not None:
            s = str(raw).strip()
            if s:
                sn_positions.append((row_idx, s))

    sn_set = {s for _, s in sn_positions}

    def is_rollup_header(s: str) -> bool:
        """True if s is an integer section or has child SNs (sub-section header)."""
        if _is_integer_section(s):
            return True
        return any(other.startswith(s + ".") for other in sn_set)

    # For each rollup header collect the statuses of its leaf questions
    # (questions that start with header + "." and are NOT themselves headers)
    section_statuses: dict[int, list] = {}
    for row_idx, sn in sn_positions:
        if not is_rollup_header(sn):
            continue
        prefix = sn + "."
        statuses = [
            review_by_sn[sub].get("status", "")
            for _, sub in sn_positions
            if sub.startswith(prefix) and not is_rollup_header(sub) and sub in review_by_sn
        ]
        section_statuses[row_idx] = statuses

    # Compute and write rollup for each header row
    for row_idx, statuses in section_statuses.items():
        if not statuses:
            continue

        total = len(statuses)
        n_app = statuses.count("APP")
        n_na  = statuses.count("NA")
        n_ir  = statuses.count("IR")

        if n_app == total:
            rollup = "APP"
        elif n_na > total / 2:
            rollup = "NA"
        elif n_ir > total / 2:
            rollup = "IR"
        else:
            rollup = "NA" if n_na >= n_ir else "IR"

        included = {"APP": "Yes", "NA": "No", "IR": "Partial"}.get(rollup, "Partial")

        inc_s = INCLUDED_STYLES[included]
        cell = ws.cell(row=row_idx, column=COL_INCL)
        cell.value = included
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, color=inc_s["fc"], bold=True)
        cell.fill = _make_fill(inc_s["bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

        st_s = STATUS_STYLES[rollup]
        cell = ws.cell(row=row_idx, column=COL_STATUS)
        cell.value = rollup
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, color=st_s["fc"], bold=True)
        cell.fill = _make_fill(st_s["bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _build_regression_comment(regression_results: list) -> str:
    """
    Build a natural-language paragraph appended to SN 6.3.6 stating the
    Python-computed regression values, whether they match reported stats,
    and whether they meet IPMVP/ASHRAE Guideline 14 thresholds.
    """
    if not regression_results:
        return ""

    blocks = []
    for res in regression_results:
        name = res.get("eem_name", "Unknown EEM")

        if res.get("error"):
            blocks.append(
                f"Independent Python regression verification ({name}) could not be completed: {res['error']}."
            )
            continue

        c          = res.get("computed") or {}
        thresholds = res.get("thresholds") or {}
        comparison = res.get("comparison") or {}

        r2        = c.get("r_squared", 0)
        cvrmse    = c.get("cv_rmse", 0)
        nmbe      = c.get("nmbe", 0)
        t_stat    = c.get("t_stat", 0)
        p_val     = c.get("p_value", 0)
        it        = c.get("intercept_t_stat")
        ip        = c.get("intercept_p_value")
        intercept = c.get("intercept", 0)
        slope     = c.get("slope", 0)
        mse       = c.get("model_std_err", 0)
        n         = c.get("n", 0)

        p_str  = "< 0.001" if p_val < 0.001 else f"= {p_val:.3f}"
        ip_str = ("< 0.001" if ip < 0.001 else f"= {ip:.4f}") if ip is not None else "N/A"
        it_str = f"{it:.4f}" if it is not None else "N/A"

        # ── Threshold split ───────────────────────────────────────────────────
        failing = [k for k, v in thresholds.items() if not v.get("passes")]
        passing = [k for k, v in thresholds.items() if v.get("passes")]

        # ── Match sentence ────────────────────────────────────────────────────
        has_reported = any(v.get("reported") is not None for v in comparison.values())
        if has_reported:
            mismatches = res.get("stats_mismatch", [])
            match_clause = (
                "consistent with those reported in the M&V Plan"
                if not mismatches else
                f"differing from the M&V Plan on: {', '.join(mismatches)}"
            )
        else:
            match_clause = "no reported statistics were available for direct comparison"

        # ── Threshold clause ──────────────────────────────────────────────────
        if not failing:
            threshold_clause = "all meet IPMVP/ASHRAE Guideline 14 thresholds."
        else:
            pass_part = f"{', '.join(passing)} {'meet' if len(passing) > 1 else 'meets'} the required thresholds" if passing else ""
            fail_part = f"{', '.join(failing)} {'do' if len(failing) > 1 else 'does'} not meet the required threshold"
            threshold_clause = f"{pass_part + '; ' if pass_part else ''}{fail_part}."

        # ── Mismatch detail sentence (only when reported stats exist) ──────────
        mismatch_detail = ""
        if has_reported and res.get("stats_mismatch"):
            details = []
            for metric in res["stats_mismatch"]:
                rep = comparison[metric].get("reported")
                comp = comparison[metric].get("computed")
                if rep is not None and comp is not None:
                    details.append(f"{metric}: reported {rep:.4g}, computed {comp:.4g}")
            if details:
                mismatch_detail = f" Discrepancies: {'; '.join(details)}."

        # ── Assemble paragraph ────────────────────────────────────────────────
        para = (
            f"[{name}] Independent Python verification (n = {n}) yields: "
            f"R² = {r2:.4f}, CV(RMSE) = {cvrmse:.2f}%, NMBE = {nmbe:.2f}%, "
            f"slope t-statistic = {t_stat:.4f} (p {p_str}), "
            f"Model Standard Error = {mse:,.2f}, intercept = {intercept:,.2f}, slope = {slope:.6f}. "
            f"These values are {match_clause}.{mismatch_detail} {threshold_clause}"
        )

        blocks.append(para)

    return "\n\n" + "\n\n".join(blocks)


def _make_calc_border() -> Border:
    s = Side(style="thin", color=CALC_BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_calc_sheet(ws, review_by_sn: dict,
                      facility_name: str = "", ref_no: str = "",
                      today_str: str = "") -> None:
    """
    Write APP/IR/NA review results into Sheet 2 or Sheet 3.
    Structure mirrors Sheet 1: data rows from row 22, col B = SN,
    cols H/I/J = Included / Active Status / Consultant's Comments.
    """
    if not review_by_sn:
        return

    # ── Header fields ─────────────────────────────────────────────────────────
    if facility_name:
        ws.cell(row=6, column=3).value = facility_name
    if ref_no:
        ws.cell(row=7, column=3).value = ref_no

    # Round 1 dates
    if today_str:
        for col in (3, 4, 5):
            ws.cell(row=15, column=col).value = today_str

    # ── Overall assessment (row 15, col F) ───────────────────────────────────
    assessment = _sheet_assessment([it.get("status", "") for it in review_by_sn.values()])
    ast_s = CALC_STATUS_STYLES[assessment]
    cell = ws.cell(row=15, column=6)
    cell.value = assessment
    cell.font = Font(name=FONT_NAME, size=CALC_FONT_SIZE, color=ast_s["fc"], bold=True)
    cell.fill = _make_fill(ast_s["bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _make_calc_border()

    # ── Data rows ─────────────────────────────────────────────────────────────
    border = _make_calc_border()
    for row_idx in range(DATA_START, ws.max_row + 1):
        sn_val = ws.cell(row=row_idx, column=COL_SN).value
        if _is_section_header(sn_val):
            continue

        sn = str(sn_val).strip()
        if sn not in review_by_sn:
            continue

        item     = review_by_sn[sn]
        included = item.get("included", "")
        status   = item.get("status", "")
        comment  = item.get("comment", "") or ""

        # Included (col H)
        inc_s = CALC_INCLUDED_STYLES.get(included, {"bg": "FFFFFF", "fc": "000000"})
        cell = ws.cell(row=row_idx, column=COL_INCL)
        cell.value = included
        cell.font = Font(name=FONT_NAME, size=CALC_FONT_SIZE, color=inc_s["fc"], bold=True)
        cell.fill = _make_fill(inc_s["bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

        # Active Status (col I)
        st_s = CALC_STATUS_STYLES.get(status, {"bg": "FFFFFF", "fc": "000000"})
        cell = ws.cell(row=row_idx, column=COL_STATUS)
        cell.value = status
        cell.font = Font(name=FONT_NAME, size=CALC_FONT_SIZE, color=st_s["fc"], bold=True)
        cell.fill = _make_fill(st_s["bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

        # Consultant's Comments (col J)
        cell = ws.cell(row=row_idx, column=COL_COMMENT)
        cell.value = comment
        if comment:
            cell.font = Font(name=FONT_NAME, size=CALC_FONT_SIZE, color=COMMENT_FC)
            cell.fill = _make_fill(COMMENT_BG)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            lines = len(comment) // 80 + comment.count("\n") + 1
            ws.row_dimensions[row_idx].height = max(30, lines * 15)
        else:
            cell.font = Font(name=FONT_NAME, size=CALC_FONT_SIZE, color=COMMENT_FC)
            cell.fill = _make_fill("FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = border

    _normalize_table_borders(ws)
    _write_section_rollups(ws, review_by_sn)


def write_review(template_bytes: bytes, review_by_sn: dict,
                 ref_no: str = "", client_name: str = "", esp_name: str = "",
                 facility_name: str = "", regression_results: list = None,
                 regression_data_provided: bool = False,
                 calc_review_sheet2: dict = None,
                 calc_review_sheet3: dict = None) -> bytes:
    """
    Load the Excel template from bytes, write review results, return as bytes.

    Parameters
    ----------
    template_bytes : bytes
        Raw bytes of the M&V Plan Review Sheet template.
    review_by_sn  : dict
        Dict keyed by SN string -> {"included": ..., "status": ..., "comment": ...}
    ref_no        : str  Reference number written to the review sheet and cover page.
    esp_name      : str  ESP name written to the cover page.
    facility_name : str  Facility name written to the review sheet.

    Returns
    -------
    bytes
        In-memory bytes of the filled workbook (xlsx).
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes), keep_links=False)

    # Drop external links, broken named ranges, and legacy drawings that
    # openpyxl cannot round-trip cleanly.
    wb._external_links.clear()

    # Clear ALL named ranges — openpyxl 3.1 cannot reliably round-trip named
    # ranges from complex templates (external refs, local sheet IDs, print
    # areas), which causes Excel to show a recovery dialog on open.  Named
    # ranges are not used by any downstream code so dropping them is safe.
    wb.defined_names.clear()
    for sheet in wb.worksheets:
        if hasattr(sheet, "defined_names") and sheet.defined_names:
            sheet.defined_names.clear()

    for sheet in wb.worksheets:
        sheet._charts.clear()
        sheet._images.clear()
        if hasattr(sheet, "_drawing") and sheet._drawing is not None:
            sheet._drawing = None

    ws = wb["1. M&V plan_V2.0"]

    # ── Fill review date fields ───────────────────────────────────────────────
    today_str = date.today().strftime("%d/%m/%y")

    # "Last Updated: dd/mm/yy" — find the cell in row 5 that contains the text
    for cell in ws[5]:
        if cell.value and "Last Updated" in str(cell.value):
            cell.value = f"Last Updated: {today_str}"
            break

    # ── Facility Name (row 6, col C) and Ref. No. (row 7, col C) ─────────────
    if facility_name:
        ws.cell(row=6, column=3).value = facility_name
    if ref_no:
        ws.cell(row=7, column=3).value = ref_no

    # ── ESP Name (row 10, col D) and Ref. No. (row 11, col D) ───────────────
    if esp_name:
        ws.cell(row=10, column=4).value = esp_name
    if ref_no:
        ws.cell(row=11, column=4).value = ref_no

    # ── Cover Page ────────────────────────────────────────────────────────────
    if "Cover Page" in wb.sheetnames:
        cp = wb["Cover Page"]
        if client_name:
            cp.cell(row=10, column=1).value = client_name
        if facility_name:
            cp.cell(row=11, column=1).value = f'Energy Efficiency Retrofit Project for "{facility_name}"'
        cp.cell(row=12, column=2).value = today_str
        if esp_name:
            cp.cell(row=13, column=2).value = esp_name
        # Date of review (col C), Open/Closed (col D), Date of Last Status (col F)
        for row in (16, 17, 18):
            cp.cell(row=row, column=3).value = today_str
            cp.cell(row=row, column=4).value = "Open"
            cp.cell(row=row, column=6).value = today_str

    # Round 1 row (row 15): Issued On (col C=3), Received on (col D=4), Reviewed on (col E=5)
    for col in (3, 4, 5):
        cell = ws.cell(row=15, column=col)
        cell.value = today_str

    # If any EEM's computed regression stats differ from those reported in the
    # M&V plan, SN 6.3.6 must be "NA" regardless of the AI verdict.
    regression_mismatch = any(
        res.get("stats_mismatch") for res in (regression_results or [])
    )

    # Round 1 Assessment (col F=6): majority-based rollup; regression mismatch forces NA.
    statuses = [item.get("status", "") for item in review_by_sn.values()]
    assessment = _sheet_assessment(statuses)
    if regression_mismatch:
        assessment = "NA"
    assess_s = STATUS_STYLES[assessment]
    _style_cell(
        ws.cell(row=15, column=6),
        assessment, bg=assess_s["bg"], fc=assess_s["fc"],
        bold=True, align="center"
    )

    # Build regression verification text once — appended to SN 6.3.6 comment
    regression_block = _build_regression_comment(regression_results or [])

    for row_idx in range(DATA_START, ws.max_row + 1):
        sn_val = ws.cell(row=row_idx, column=COL_SN).value
        if _is_section_header(sn_val):
            continue

        sn = str(sn_val).strip()
        if sn not in review_by_sn:
            continue

        item     = review_by_sn[sn]
        included = item.get("included", "")
        status   = item.get("status", "")
        comment  = item.get("comment", "") or ""

        # If computed regression stats differ from reported, force NA
        if sn == "6.3.6" and regression_mismatch:
            status = "NA"

        # Append Python regression verification to the 6.3.6 comment
        if sn == "6.3.6":
            if regression_block:
                comment = (comment + regression_block).strip()
            elif included in ("Yes", "Partial") and not regression_data_provided:
                not_provided = (
                    "\n\nRegression data was not provided; independent Python "
                    "verification of the reported regression statistics could not be performed."
                )
                comment = (comment + not_provided).strip()

        # ── Included (col H) ────────────────────────────────────────────────
        inc_s = INCLUDED_STYLES.get(included, {"bg": "FFFFFF", "fc": "000000"})
        _style_cell(
            ws.cell(row=row_idx, column=COL_INCL),
            included, bg=inc_s["bg"], fc=inc_s["fc"],
            bold=True, align="center"
        )

        # ── Active Status (col I) ────────────────────────────────────────────
        st_s = STATUS_STYLES.get(status, {"bg": "FFFFFF", "fc": "000000"})
        _style_cell(
            ws.cell(row=row_idx, column=COL_STATUS),
            status, bg=st_s["bg"], fc=st_s["fc"],
            bold=True, align="center"
        )

        # ── Consultant's Comment (col J) ─────────────────────────────────────
        if comment:
            _style_cell(
                ws.cell(row=row_idx, column=COL_COMMENT),
                comment, bg=COMMENT_BG, fc=COMMENT_FC,
                wrap=True, align="left"
            )
            lines = len(comment) // 80 + comment.count("\n") + 1
            ws.row_dimensions[row_idx].height = max(40, lines * 15)
        else:
            _style_cell(
                ws.cell(row=row_idx, column=COL_COMMENT),
                "", bg="FFFFFF", fc=COMMENT_FC,
                wrap=True, align="left"
            )

    # ── Sheet 1 — Consistent thin borders then section rollups ───────────────
    _normalize_table_borders(ws)
    _write_section_rollups(ws, review_by_sn)

    # ── Sheet 1 — Active Status dropdown (col I, data rows) ─────────────────
    # Remove any existing validation on col I so the new one takes precedence.
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if not any(str(r).startswith("I") for r in dv.sqref.ranges)
    ]
    dv_status = DataValidation(
        type="list",
        formula1='"APP,AAN,NA,IR"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_status.sqref = f"I{DATA_START}:I{ws.max_row + 50}"
    ws.add_data_validation(dv_status)

    # ── Sheets 2 & 3 — M&V Calculations and Sample M&V Reports ──────────────
    if calc_review_sheet2 and "2. M&V Calculations" in wb.sheetnames:
        _write_calc_sheet(
            wb["2. M&V Calculations"], calc_review_sheet2,
            facility_name=facility_name, ref_no=ref_no, today_str=today_str,
        )
    if calc_review_sheet3 and "3. Sample M&V Reports" in wb.sheetnames:
        _write_calc_sheet(
            wb["3. Sample M&V Reports"], calc_review_sheet3,
            facility_name=facility_name, ref_no=ref_no, today_str=today_str,
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
