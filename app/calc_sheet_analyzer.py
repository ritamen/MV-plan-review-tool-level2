"""
calc_sheet_analyzer.py
----------------------
Python-only structural analysis of the M&V Calculation Sheet Excel workbook.
Answers all 9 Sheet 2 questions without calling the AI.

Strategy:
  1. Extract every sheet's name + cell text (first MAX_SCAN_ROWS rows).
  2. Match keyword sets to determine whether each required section exists.
  3. Return a dict keyed by SN -> {included, status, comment}.

Status values: APP / IR / NA  (consistent with Sheet 2 colour coding)
"""

import io
import logging
from typing import Optional

import openpyxl

MAX_SCAN_ROWS = 60   # rows scanned per sheet for keyword evidence


# ── Helpers ───────────────────────────────────────────────────────────────────

def _sheet_text(ws) -> str:
    """Return all cell values from the first MAX_SCAN_ROWS rows as a single lowercased string."""
    parts = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > MAX_SCAN_ROWS:
            break
        for cell in row:
            if cell is not None:
                parts.append(str(cell).strip().lower())
    return " ".join(parts)


def _has_keywords(text: str, keywords: list[str], min_matches: int = 1) -> bool:
    return sum(1 for kw in keywords if kw.lower() in text) >= min_matches


def _count_keywords(text: str, keywords: list[str]) -> int:
    return sum(1 for kw in keywords if kw.lower() in text)


def _has_numeric_data(ws, min_rows: int = 5) -> bool:
    """Return True if the sheet has at least min_rows rows with numeric values."""
    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(isinstance(c, (int, float)) for c in row):
            count += 1
            if count >= min_rows:
                return True
    return False


# ── Main analyzer ─────────────────────────────────────────────────────────────

def analyze_calc_sheet(excel_bytes: bytes) -> dict:
    """
    Analyze the ESP's M&V Calculation Sheet workbook and return Sheet 2 results.

    Returns dict: SN -> {"included": "Yes"|"No"|"Partial", "status": "APP"|"IR"|"NA", "comment": str}
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    sheet_names = wb.sheetnames

    # Build per-sheet text fingerprints and numeric-data flags
    sheets: list[dict] = []
    for name in sheet_names:
        ws = wb[name]
        text = _sheet_text(ws)
        sheets.append({
            "name":    name,
            "name_lc": name.lower(),
            "text":   text,
            "has_nums": _has_numeric_data(ws),
            "ws":     ws,
        })

    wb.close()

    results = {}

    # ── 0.1  Introductory / cover sheet (must be the FIRST sheet) ─────────────
    intro_keywords = ["m&v plan", "reporting period", "esp", "target savings", "facility", "client"]
    if sheets:
        first = sheets[0]
        hits = _count_keywords(first["text"], intro_keywords)
        name = first["name"]
        if hits >= 3:
            results["0.1"] = {
                "included": "Yes", "status": "APP",
                "comment": f"Confirmed. First sheet ('{name}') contains references to: "
                           f"{', '.join(kw for kw in intro_keywords if kw in first['text'])}.",
            }
        elif hits >= 1:
            results["0.1"] = {
                "included": "Partial", "status": "IR",
                "comment": f"Sheet '{name}' is present as the first sheet but appears to be missing "
                           f"some required fields (found: {hits}/{len(intro_keywords)} expected items). "
                           f"ESP shall ensure the introductory sheet references the M&V Plan, reporting "
                           f"period dates, ESP name, target savings, facility name, and client name.",
            }
        else:
            results["0.1"] = {
                "included": "No", "status": "NA",
                "comment": "No introductory cover sheet was identified as the first sheet of the workbook. "
                           "ESP shall include a dedicated introductory sheet as the first tab.",
            }
    else:
        results["0.1"] = {
            "included": "No", "status": "NA",
            "comment": "The uploaded workbook appears to contain no sheets.",
        }

    # ── 1.1  Main calculation sheet ────────────────────────────────────────────
    main_calc_keywords = ["baseline", "reporting period", "savings", "adjustment", "target savings"]
    best = max(sheets, key=lambda s: _count_keywords(s["text"], main_calc_keywords), default=None)
    if best and _count_keywords(best["text"], main_calc_keywords) >= 3:
        results["1.1"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Main calculation sheet identified ('{best['name']}') containing "
                       f"baseline data, reporting period data, savings calculations, and adjustment references.",
        }
    elif best and _count_keywords(best["text"], main_calc_keywords) >= 1:
        results["1.1"] = {
            "included": "Partial", "status": "IR",
            "comment": f"A calculation sheet was found ('{best['name']}') but it appears incomplete. "
                       f"ESP shall include a dedicated main calculation sheet covering baseline data, "
                       f"reporting period data, routine and non-routine adjustments, savings calculations, "
                       f"and comparison with target savings.",
        }
    else:
        results["1.1"] = {
            "included": "No", "status": "NA",
            "comment": "No main calculation sheet was identified. ESP shall include a dedicated sheet "
                       "showing baseline data, reporting period data, adjustments, savings calculations, "
                       "and comparison with target savings.",
        }

    # ── 2.1  Regression model sheet ────────────────────────────────────────────
    reg_keywords = ["regression", "r²", "r2", "r-squared", "r square", "slope", "intercept",
                    "coefficient", "correlation", "ols"]
    reg_matches = [s for s in sheets if _has_keywords(s["text"] + " " + s["name_lc"], reg_keywords)]
    if reg_matches:
        names = ", ".join(f"'{s['name']}'" for s in reg_matches)
        results["2.1"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Regression model sheet(s) identified: {names}.",
        }
    else:
        results["2.1"] = {
            "included": "No", "status": "NA",
            "comment": "No regression model sheet was identified. If regression is used as the routine "
                       "adjustment model, ESP shall include a dedicated regression sheet per the M&V Plan.",
        }

    # ── 3.1  15-minute interval data sheet ─────────────────────────────────────
    interval_keywords = ["15-min", "15 min", "15min", "interval", "15-minute", "15 minute"]
    supporting_keywords = ["kwh", "kw", "energy", "demand", "power", "consumption"]
    interval_matches = [
        s for s in sheets
        if _has_keywords(s["text"] + " " + s["name_lc"], interval_keywords)
        or (_has_keywords(s["text"] + " " + s["name_lc"], supporting_keywords) and s["has_nums"]
            and "15" in s["text"])
    ]
    if interval_matches:
        names = ", ".join(f"'{s['name']}'" for s in interval_matches)
        results["3.1"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. 15-minute interval data sheet(s) identified: {names}.",
        }
    else:
        results["3.1"] = {
            "included": "No", "status": "NA",
            "comment": "No 15-minute interval data sheet was identified. ESP shall include a sample "
                       "sheet of extracted 15-minute interval data used to determine energy consumption "
                       "and key parameters.",
        }

    # ── 3.2  Utility bills / solar data sheet ──────────────────────────────────
    utility_keywords = ["utility", "bill", "invoice", "solar", "tariff", "electricity bill",
                        "dewa", "addc", "sewa", "fewa", "aadc", "meter reading"]
    utility_matches = [s for s in sheets if _has_keywords(s["text"] + " " + s["name_lc"], utility_keywords)]
    if utility_matches:
        names = ", ".join(f"'{s['name']}'" for s in utility_matches)
        results["3.2"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Utility bills/solar data sheet(s) identified: {names}.",
        }
    else:
        results["3.2"] = {
            "included": "No", "status": "NA",
            "comment": "No utility bills or solar data sheet was identified. ESP shall include a sample "
                       "sheet showing tabulated utility bills and solar data, if applicable.",
        }

    # ── 4.1  Routine adjustment model sheet (non-regression) ───────────────────
    # Only expected if regression is NOT the main model.
    routine_keywords = ["routine adjustment", "routine", "independent variable", "adjustment model",
                        "cdd", "hdd", "occupancy", "temperature", "driving variable"]
    nra_keywords = ["non-routine", "non routine", "nra"]
    routine_matches = [
        s for s in sheets
        if _has_keywords(s["text"] + " " + s["name_lc"], routine_keywords)
        and not _has_keywords(s["text"], nra_keywords)   # exclude NRA sheets
        and not _has_keywords(s["text"] + " " + s["name_lc"], reg_keywords)  # exclude regression sheets
    ]
    if routine_matches:
        names = ", ".join(f"'{s['name']}'" for s in routine_matches)
        results["4.1"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Routine adjustment model sheet(s) identified: {names}.",
        }
    elif reg_matches:
        # Regression is used — this question may not apply
        results["4.1"] = {
            "included": "Partial", "status": "IR",
            "comment": "Regression appears to be the primary routine adjustment model. If a non-regression "
                       "routine adjustment model is also used, ESP shall include a dedicated sheet with the "
                       "model and a sample table of independent variables.",
        }
    else:
        results["4.1"] = {
            "included": "No", "status": "NA",
            "comment": "No routine adjustment model sheet was identified. ESP shall present a sample of "
                       "the routine adjustment model and a sample table of independent variables to be captured.",
        }

    # ── 5.1 & 5.2  Non-routine adjustment sheets ───────────────────────────────
    nra_matches = [s for s in sheets if _has_keywords(s["text"] + " " + s["name_lc"], nra_keywords)]

    if len(nra_matches) >= 2:
        # Two or more NRA sheets — assign: first = calculations, second = events list
        s1, s2 = nra_matches[0], nra_matches[1]
        results["5.1"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Non-routine adjustments sheet identified: '{s1['name']}'.",
        }
        results["5.2"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Non-routine events list sheet identified: '{s2['name']}'.",
        }
    elif len(nra_matches) == 1:
        s = nra_matches[0]
        event_keywords = ["event", "date", "description", "list"]
        has_event_list = _has_keywords(s["text"], event_keywords)
        results["5.1"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Non-routine adjustments sheet identified: '{s['name']}'.",
        }
        results["5.2"] = {
            "included": "Partial", "status": "IR",
            "comment": f"Only one non-routine sheet found ('{s['name']}'). ESP shall include a separate "
                       f"sheet listing all non-routine events (date, description, and impact).",
        }
    else:
        results["5.1"] = {
            "included": "No", "status": "NA",
            "comment": "No non-routine adjustments sheet was identified. ESP shall include a dedicated "
                       "sheet for non-routine adjustment calculations.",
        }
        results["5.2"] = {
            "included": "No", "status": "NA",
            "comment": "No non-routine events list sheet was identified. ESP shall include a dedicated "
                       "sheet listing all non-routine events.",
        }

    # ── 6.1  Uncertainty quantification sheet ──────────────────────────────────
    uncert_keywords = ["uncertainty", "confidence interval", "confidence level", "error analysis",
                       "measurement uncertainty", "tolerance", "uncertainty analysis"]
    uncert_matches = [s for s in sheets if _has_keywords(s["text"] + " " + s["name_lc"], uncert_keywords)]
    if uncert_matches:
        names = ", ".join(f"'{s['name']}'" for s in uncert_matches)
        results["6.1"] = {
            "included": "Yes", "status": "APP",
            "comment": f"Confirmed. Uncertainty quantification sheet(s) identified: {names}.",
        }
    else:
        results["6.1"] = {
            "included": "No", "status": "NA",
            "comment": "No uncertainty quantification sheet was identified. ESP shall include a dedicated "
                       "sheet quantifying uncertainty issues.",
        }

    logging.info(
        "calc_sheet_analyzer: %d sheets scanned, %d questions answered.",
        len(sheets), len(results),
    )
    return results
